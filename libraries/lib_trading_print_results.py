__author__ = "Luis Domingues"


#----------------------------------------------------------------------------------------
# Notes
#----------------------------------------------------------------------------------------


#----------------------------------------------------------------------------------------
# IMPORTS
#----------------------------------------------------------------------------------------
import lib_excel_ops_openpyxl as lib_excel
import lib_general_ops as lib_gen
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from trading_classes import Stock


#----------------------------------------------------------------------------------------
# INPUTS
#----------------------------------------------------------------------------------------


#----------------------------------------------------------------------------------------
# FUNCTIONS
#----------------------------------------------------------------------------------------
def write_results_to_excel(params_dic, wb, mngr_table=None, write_params=True, write_table=True):
    # Get worksheet for parameters and execution/backtesting/optimization
    if write_params:
        ws_params = lib_excel.get_worksheet(wb, 'Parameters')
        assert ws_params != None
    ws_name = None
    if write_table:
        if params_dic['Mode'] == 'Analysis':
            ws_name = 'Execution'
        if params_dic['Mode'] == 'Backtesting':
            ws_name = 'Backtesting'
        if params_dic['Mode'] == 'Optimization':
            ws_name = 'Optimization'
        ws = lib_excel.get_worksheet(wb, ws_name)
        assert ws != None

    # Write parameters
    if write_params:
        write_parameters(params_dic, ws_params)
    # Write results for execution/backtesting
    if write_table:
        write_data(params_dic, mngr_table, ws)
    return 0

def write_parameters(params_dic, ws_params):
    # Write parameters
    row_i = 8
    ws_params.cell(row=3, column=2, value=params_dic['Algorithm name'])
    for param in params_dic.keys():
        ws_params.cell(row=row_i, column=1, value=param)
        try:
            ws_params.cell(row=row_i, column=2, value=params_dic[param])
        except:
            ws_params.cell(row=row_i, column=2, value=str(params_dic[param]))
        row_i = row_i + 1
    return 0

def write_data(params_dic, mngr_table, ws):
    ft = Font(name='Calibri', size=11, bold=True)
    al = Alignment(horizontal='center', vertical='center', wrap_text=True)
    # Write results
    if params_dic['Mode'] == 'Analysis':
        ws.cell(row=3, column=2, value=params_dic['Algorithm name'])
        # Write orders
        orders_list = mngr_table.get_orders_list()
        row_i = lib_excel.determine_first_empty_row(ws, row_start=8)
        for order in orders_list:
            st = order.get_stock()
            id = row_i - 8 + 1
            ws.cell(row=row_i, column=1, value=id)
            c.alignment = Alignment(horizontal='center')
            ws.cell(row=row_i, column=2, value=order.get_name())
            c.alignment = Alignment(horizontal='left')
            ws.cell(row=row_i, column=3, value=order.get_position_type())
            c.alignment = Alignment(horizontal='center')
            ws.cell(row=row_i, column=4, value=order.get_order_type())
            c.alignment = Alignment(horizontal='center')
            ws.cell(row=row_i, column=5, value=order.get_amount())
            c.number_format = '0'
            c.alignment = Alignment(horizontal='right')
            ws.cell(row=row_i, column=6, value=order.get_stop_loss())
            c.number_format = '0.0000'
            c.alignment = Alignment(horizontal='right')
            ws.cell(row=row_i, column=7, value=order.get_date())
            c.alignment = Alignment(horizontal='center')
            ws.cell(row=row_i, column=8, value=order.get_price())
            c.number_format = '0.00'
            c.alignment = Alignment(horizontal='right')
            i = 0
            c_dic = order.get_custom()
            for c in c_dic.keys():
                ws.cell(row=6, column=9+i, value='python')
                w = ws.cell(row=7, column=9+i, value=c)
                w.font = ft
                w.alignment = al
                w = ws.cell(row=row_i, column=9+i, value=c_dic[c])
                if lib_gen.is_number(c_dic[c]):
                    w.number_format = '0.000000'
                    w.alignment = Alignment(horizontal='right')
                else:
                    w.alignment = Alignment(horizontal='center')
                i += 1
            row_i = row_i + 1
        #xl_tab = Table(displayName='Tbl', ref="A7:L13")
        #style = TableStyleInfo(name='TableStyleMedium9', showFirstColumn=False,
        #                       showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        #xl_tab.tableStyleInfo = style
        #ws.add_table(xl_tab)

    if params_dic['Mode'] == 'Backtesting':
        for k in mngr_table.keys():
            positions_list = mngr_table[k].get_positions_list()
            row_i = lib_excel.determine_first_empty_row(ws, row_start=8)
            ws.cell(row=3, column=2, value=params_dic['Algorithm name'])
            ws.cell(row=4, column=2, value=mngr_table[k].get_profit_losses_pct())
            for position in positions_list:
                id = row_i - 8 + 1
                c = ws.cell(row=row_i, column=1, value=id)
                c.alignment = Alignment(horizontal='center')
                c = ws.cell(row=row_i, column=2, value=position.get_name())
                c.alignment = Alignment(horizontal='left')
                c = ws.cell(row=row_i, column=3, value=position.get_position_type())
                c.alignment = Alignment(horizontal='center')
                c = ws.cell(row=row_i, column=4, value=position.get_state())
                c.alignment = Alignment(horizontal='center')
                c = ws.cell(row=row_i, column=5, value=position.get_amount())
                c.number_format = '0'
                c.alignment = Alignment(horizontal='center')
                c = ws.cell(row=row_i, column=6, value=position.get_stop_loss())
                c.number_format = '0.0000'
                c.alignment = Alignment(horizontal='right')
                c = ws.cell(row=row_i, column=7, value=position.get_date_entry())
                c.alignment = Alignment(horizontal='center')
                c = ws.cell(row=row_i, column=8, value=position.get_price_entry())
                c.number_format = '0.00'
                c.alignment = Alignment(horizontal='right')
                i = 0
                c_dic = position.get_custom_entry()
                for c in c_dic.keys():
                    ws.cell(row=6, column=15+i, value='python')
                    w = ws.cell(row=7, column=15+i, value=c)
                    w.font = ft
                    w.alignment = al
                    w = ws.cell(row=row_i, column=15+i, value=c_dic[c])
                    if lib_gen.is_number(c_dic[c]):
                        w.number_format = '0.0000'
                        w.alignment = Alignment(horizontal='right')
                    else:
                        w.alignment = Alignment(horizontal='center')
                    i += 1
                if position.get_state() == 'closed':
                    c = ws.cell(row=row_i, column=9, value=position.get_date_exit())
                    c.alignment = Alignment(horizontal='center')
                    c = ws.cell(row=row_i, column=10, value=position.get_price_exit())
                    c.number_format = '0.00'
                    c.alignment = Alignment(horizontal='right')
                    c = ws.cell(row=row_i, column=11, value=position.get_money_entry())
                    c.number_format = '0.00'
                    c.alignment = Alignment(horizontal='right')
                    c = ws.cell(row=row_i, column=12, value=position.get_money_exit())
                    c.number_format = '0.00'
                    c.alignment = Alignment(horizontal='right')
                    c = ws.cell(row=row_i, column=13, value=position.get_profit_losses())
                    c.number_format = '0.00'
                    c.alignment = Alignment(horizontal='right')
                    c = ws.cell(row=row_i, column=14, value=position.get_profit_losses_pct())
                    c.number_format = '0.0'
                    c.alignment = Alignment(horizontal='right')
                    c_dic = position.get_custom_exit()
                    for c in c_dic.keys():
                        ws.cell(row=6, column=15 + i, value='python')
                        w = ws.cell(row=7, column=15 + i, value=c)
                        w.font = ft
                        w.alignment = al
                        w = ws.cell(row=row_i, column=15 + i, value=c_dic[c])
                        if lib_gen.is_number(c_dic[c]):
                            w.number_format = '0.000000'
                            w.alignment = Alignment(horizontal='right')
                        else:
                            w.alignment = Alignment(horizontal='center')
                        i += 1
                row_i = row_i + 1

    if params_dic['Mode'] == 'Optimization':
        pass

    return 0
