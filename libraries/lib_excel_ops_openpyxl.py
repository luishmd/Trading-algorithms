"""
Library for excel operations using openpyxl
"""

import openpyxl as xl
from openpyxl.utils import get_column_letter, coordinate_from_string, column_index_from_string


# Metadata
__author__="Luis Domingues"
__copyright__ = "Copyright 2017, Process Systems Enterprise, Ltd."
__version__ = "1.0.0"
__maintainer__ = "Luis Domingues"
__email__ = "l.domingues@psenterprise.com"
__status__ = "Development"


#----------------------------------------------------------------------------------------
# FUNCTIONS
#----------------------------------------------------------------------------------------
def create_workbook():
    """
    Function that returns an empty excel workbook object
    """
    return xl.Workbook()


def create_worksheet(wb, title, index=None):
    """
    Function that creates a worksheet and returns both the wb and the ws. If no index is specified the worksheet is added to the end. If the worksheet already exists it returns it.
    """
    if title in get_worksheet_names(wb):
        ws = get_worksheet(wb, title)
    else:
        ws = wb.create_sheet(title=title, index=index)
    return ws


def open_workbook(path_to_workbook_file):
    """
    Function that returns loads a workbook and returns it as an object
    """
    try:
        wb = xl.load_workbook(path_to_workbook_file)
        return wb
    except:
        print("Could not load workbook <%s>" % path_to_workbook_file)
        return None


def convert_coordinates_num_to_alphanum(numeric_coordinates_list, zero_indexed=False):
    """
    Function that converts coordinates from a numeric row and index to alphanumeric (e.g. [1,1] -> 'A1')
    """
    [row, column] = numeric_coordinates_list
    if zero_indexed:
        row += 1
        column += 1
    return get_column_letter(column) + str(row)


def convert_coordinates_alphanum_to_num(alphanum_coordinates_str, zero_indexed=False):
    """
    Function that converts coordinates from a alphanumeric inde to a numeric one (e.g. 'A1' -> [1,1])
    """
    xy = coordinate_from_string(alphanum_coordinates_str) # returns ('A',1)
    if zero_indexed:
        col = column_index_from_string(xy[0]) - 1
        row = xy[1] - 1
    else:
        col = column_index_from_string(xy[0])
        row = xy[1]
    return [row, col]


def convert_column_num_to_str(col_num, zero_indexed=False):
    """
    Function that converts a column from a number to a string (e.g. 1 -> 'A')
    """
    if zero_indexed:
        col_str = get_column_letter(col_num+1)
    else:
        col_str = get_column_letter(col_num)
    return col_str


def convert_column_str_to_num(col_str, zero_indexed=False):
    """
    Function that converts a column from a string to a number (e.g. 'A' -> 1)
    """
    if zero_indexed:
        col_num = column_index_from_string(col_str) - 1
    else:
        col_num = column_index_from_string(col_str)
    return col_num


def get_worksheet(wb, ws_name):
    """
    Function that returns the worksheet with name ws_name
    """
    try:
        ws = wb[ws_name]
        return ws
    except:
        print("Worksheet %s does no exist in workbook." % ws_name)
        return None


def get_worksheet_names(wb):
    """
    Function that returns a list containing all worksheet names
    """
    worksheet_names_list = wb.sheetnames
    return worksheet_names_list


def write_label(ws, label_str, coordinates_str=None, coordinates_list=None, zero_indexed=False):
    """
    Function that writes a label in a worksheet
    """
    if coordinates_str != None:
        ws[coordinates_str] = label_str
    if coordinates_list != None:
        coordinates = convert_coordinates_num_to_alphanum(coordinates_list, zero_indexed=zero_indexed)
        ws[coordinates] = label_str
    return 0


def write_data_array(ws, data, coordinates_str=None, coordinates_list=None, orientation='columns', zero_indexed=False):
    """
    Function that writes an array of data in a worksheet
    """
    if coordinates_str != None:
        coordinates_list = convert_coordinates_alphanum_to_num(coordinates_str, zero_indexed=zero_indexed)
    if zero_indexed:
        row = coordinates_list[0] + 1
        col = coordinates_list[1] + 1
    else:
        row = coordinates_list[0]
        col = coordinates_list[1]
    if type(data) == type([]): # if it is a list it must be an array (no list of lists)
        length = len(data)
        if orientation.lower() == "columns":
            for i in range(0,length):
                ws.cell(row=row+i, column=col).value = data[i]
        if orientation.lower() == "rows":
            for i in range(0, length):
                ws.cell(row=row, column=col+i).value = data[i]
    else: # scalar
        ws.cell(row=row, column=col).value = data
    return 0


def write_data_matrix(ws, data, coordinates_str=None, coordinates_list=None, orientation='columns', zero_indexed=False):
    """
    Function that writes a matrix of data in a worksheet
    """
    if coordinates_str != None:
        coordinates_list = convert_coordinates_alphanum_to_num(coordinates_str, zero_indexed=zero_indexed)
    if zero_indexed:
        row = coordinates_list[0] + 1
        col = coordinates_list[1] + 1
    else:
        row = coordinates_list[0]
        col = coordinates_list[1]
    data_is_matrix = False
    try: # Check if data is matrix (list of lists)
        aux = data[0]
        if type(aux) == type([]) and type(aux[0]) != type([]):
            data_is_matrix = True
        assert data_is_matrix
    except:
        print("Data to write is not a matrix")
    if data_is_matrix:
        if orientation.lower() == "columns":
            for i in range(0,len(data)):
                data_array = data[i]
                coordinates_list = [row, col + i]
                write_data_array(ws, data_array, coordinates_str=coordinates_str, coordinates_list=coordinates_list, orientation='columns', zero_indexed=zero_indexed)
        if orientation.lower() == "rows":
            for i in range(0,len(data)):
                data_array = data[i]
                coordinates_list = [row + i, col]
                write_data_array(ws, data_array, coordinates_str=coordinates_str, coordinates_list=coordinates_list, orientation='rows', zero_indexed=zero_indexed)
    return 0


def save_workbook(wb, file_complete_path):
    """
    Function that saves an excel workbook in path_to_save
    """
    try:
        wb.save(file_complete_path)
        return 0
    except:
        print("Could not save workbook to directory <%s>" % path_to_save)
        return 1


def determine_first_empty_row(ws, row_start=1):
    """
    Function that determines first empty row and resturns its index
    """
    row_i = row_start
    while ws.cell(row=row_i, column=1).value != None:
        row_i = row_i + 1
    return row_i

#----------------------------------------------------------------------------------------
# MAIN
#----------------------------------------------------------------------------------------
if __name__ == "__main__":
    pass

    # Testing
    #wb = create_workbook()
    #path_to_save = "C:/Users/luisd/Desktop/test.xlsx"
    #print convert_coordinates_num_to_alphanum([1,1], zero_indexed=False)
    #print convert_coordinates_alphanum_to_num('A1', zero_indexed=False)
    #print convert_column_str_to_num('C', zero_indexed=False)
    #print convert_column_num_to_str(3, zero_indexed=False)
    #ws = create_worksheet(wb, "name")
    #ws = get_worksheet(wb, "name")
    #data = [1,2,3]
    #data = 5
    #write_data_array(ws, data, coordinates_list=[1,1], orientation = "columns")
    #save_workbook(wb, path_to_save)
    #print get_worksheet_names(wb)

    #wb = create_workbook()
    #save_workbook(wb, path_to_save)